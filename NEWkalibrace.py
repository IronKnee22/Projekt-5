import time
import tkinter
import tkinter.messagebox
import openpyxl
import customtkinter

from PneumoCVUTFBMI.DeviceLoader import DeviceLoader

workbook = openpyxl.Workbook()
worksheet = workbook.active
Minuly_Sval = 0
Zvoleny_Sval = 0
y = 0
x = 1

dl=DeviceLoader()

b1 = dl.getBoard1()
b2 = dl.getBoard2()
b3 = dl.getBoard3()
b4 = dl.getBoard4()
b5 = dl.getBoard5()

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Kalibrační GUI")
        self.geometry(f"{540}x{220}")

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="Výběr svalu", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        # create radiobutton frame
        self.radio_var = tkinter.IntVar(value=0)
        
        self.radio_button_1 = customtkinter.CTkRadioButton(master=self.sidebar_frame, variable=self.radio_var, value=1, command=self.radiobutton_event, text="Sval 1")
        self.radio_button_1.grid(row=1, column=0)
        self.radio_button_2 = customtkinter.CTkRadioButton(master=self.sidebar_frame, variable=self.radio_var, value=2, command=self.radiobutton_event, text="Sval 2")
        self.radio_button_2.grid(row=2, column=0)
        self.radio_button_3 = customtkinter.CTkRadioButton(master=self.sidebar_frame, variable=self.radio_var, value=3, command=self.radiobutton_event, text="Sval 3")
        self.radio_button_3.grid(row=3, column=0)
        self.radio_button_4 = customtkinter.CTkRadioButton(master=self.sidebar_frame, variable=self.radio_var, value=4, command=self.radiobutton_event, text="Sval 4")
        self.radio_button_4.grid(row=4, column=0)
        self.radio_button_5 = customtkinter.CTkRadioButton(master=self.sidebar_frame, variable=self.radio_var, value=5, command=self.radiobutton_event, text="Sval 5")
        self.radio_button_5.grid(row=5, column=0)

        #Apearence mode options
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 10))
        self.appearance_mode_optionemenu.set("System")

         # create main entry and button
        self.entrySpeed = customtkinter.CTkEntry(self, placeholder_text="Speed")
        self.entrySpeed.grid(row=2, column=1,  padx=(20, 0), pady=(20, 20), sticky="nsew")

        self.entrySteps = customtkinter.CTkEntry(self, placeholder_text="Steps")
        self.entrySteps.grid(row=2, column=2,padx=(20, 0), pady=(20, 20), sticky="nsew")

        self.entry = customtkinter.CTkEntry(self, placeholder_text="Hw Value")
        self.entry.grid(row=3, column=1, padx=(20, 0), pady=(20, 20), sticky="nsew")

        self.main_button_1 = customtkinter.CTkButton(master=self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.hw_value)
        self.main_button_1.grid(row=3, column=2, padx=(20, 20), pady=(20, 20), sticky="nsew")
        self.main_button_1.bind('<Return>', self.hw_value, add='+')# tohle nevím proč nefunguje je to funkce na přidávání pomocí enteru

        
    

    def hw_value(self):
        global x
        global y
        global Speed
        global Steps
        global Minuly_Sval

        # Vytvoření slovníku pro objekty b1, b2, atd.
        global b_objects
        b_objects = {
            1: b1,
            2: b2,
            3: b3,
            4: b4,
            5: b5
        }

        if Minuly_Sval != Zvoleny_Sval:
            if Minuly_Sval in b_objects:
                for i in range(1, x):
                    b_objects[Minuly_Sval].go_backward(Steps, Speed)
                    time.sleep(3)

            worksheet.cell(row=1, column=1, value="Počet krokůmotoru") 
            worksheet.cell(row=1, column=2, value="Hodnota v mV") 
            worksheet.cell(row=1, column=3, value="Hodnota fyzického senzoru") 
            worksheet.cell(row=1, column=4, value="Rychlost") 
            workbook.save('sval' + str(Minuly_Sval)+'.xlsx')
            print("změna svalu")
            x = 1
            y = 0


     
        
        
        if Zvoleny_Sval in b_objects:
            b_object = b_objects[Zvoleny_Sval]

            b_object.on()
            Speed = self.entrySpeed.get()
            Steps = self.entrySteps.get()
            y = y + int(Steps)  # Tady se nám počítá kolik udělal motor celkově kroků
            x = x + 1
            b_object.go_forward(Steps, Speed)
            time.sleep(3)
            worksheet.cell(row=x, column=1, value=y)  # počet kroků motoru
            worksheet.cell(row=x, column=2, value=str(b_object.readA0()))  # Zde se načítá ze senzoru může být problém
            worksheet.cell(row=x, column=3, value=self.entry.get())  # hodnota v mv
            worksheet.cell(row=x, column=4, value=Speed)  # hodnota Rychlosti
            Minuly_Sval = Zvoleny_Sval
            self.entry.delete(0, 10000)
            print(x)
        else:
            print("Neplatný Zvoleny_Sval")

    def open_input_dialog_event(self):
        dialog = customtkinter.CTkInputDialog(text="Type in a number:", title="CTkInputDialog")
        print("CTkInputDialog:", dialog.get_input())

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def getRadiVar(self):
        print("Počáteční hodnota:", self.radio_var.get())
        

    def radiobutton_event(self):
        global Zvoleny_Sval
        
        selected_value = self.radio_var.get()

        if 1 <= selected_value <= 5:
            print(f"Sval {selected_value}")
            Zvoleny_Sval = selected_value
        else:
            print("Neznámý sval")

if __name__ == "__main__":
    app = App()
    app.getRadiVar()
    app.mainloop()

worksheet.cell(row=1, column=1, value="Počet krokůmotoru") #počet kroků motoru
worksheet.cell(row=1, column=2, value="Hodnota v mV") #hodnota v mv
worksheet.cell(row=1, column=3, value="Hodnota fyzického senzoru") #Zde se načítá ze senzoru může být problém
worksheet.cell(row=1, column=4, value="Rychlost") #hodnota Rychlosti
workbook.save('sval' + str(Zvoleny_Sval)+'.xlsx')


if Minuly_Sval in b_objects:
    for i in range(1, x):
        b_objects[Minuly_Sval].go_backward(Steps, Speed)
        time.sleep(3)
